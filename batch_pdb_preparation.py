#!/usr/bin/env python3
"""
Batch PDB Preparation Module
============================

Handles batch processing of multiple PDB files for molecular docking preparation
using YAML or TXT configuration files.

Author: Molecular Docking Pipeline
Version: 1.0.0
"""

import argparse
import sys
import json
import yaml
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd

# Import the core pipeline
from core_pipeline import MolecularDockingPipeline, extract_residue_level_coordinates, EXCEL_AVAILABLE
from autodock_preparation import AutoDockPreparationPipeline, PreparationConfig

if EXCEL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment


class BatchPDBPreparationPipeline:
    """
    Batch processing pipeline for multiple PDB files with configuration support
    """
    
    def __init__(self, config_file: str, output_dir: str = "batch_docking_preparation"):
        """
        Initialize the batch processing pipeline
        
        Args:
            config_file: Path to configuration file (YAML or TXT)
            output_dir: Base output directory for all results
        """
        self.config_file = Path(config_file)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.config = self._load_config()
        self.results = {}
        
    def _load_config(self) -> Dict[str, Any]:
        """
        Load configuration from YAML or TXT file
        
        Returns:
            Dictionary with configuration data
        """
        if not self.config_file.exists():
            raise FileNotFoundError(f"Configuration file not found: {self.config_file}")
            
        if self.config_file.suffix.lower() == '.yaml' or self.config_file.suffix.lower() == '.yml':
            with open(self.config_file, 'r') as f:
                return yaml.safe_load(f)
        elif self.config_file.suffix.lower() == '.txt':
            return self._load_txt_config()
        else:
            raise ValueError(f"Unsupported configuration file format: {self.config_file.suffix}")
            
    def _load_txt_config(self) -> Dict[str, Any]:
        """
        Load configuration from TXT file
        
        Returns:
            Dictionary with configuration data
        """
        config = {
            "global_settings": {
                "output_directory": str(self.output_dir),
                "cleaning": {
                    "strategy": "common",
                    "common_residues": ["HOH", "NA", "CL", "SO4", "CA", "MG", "ZN", "FE", "CU", "MN"]
                },
                "analysis": {
                    "use_enhanced_coordinates": True,
                    "distance_cutoff": 5.0,
                    "method": "plip"
                },
                "autodock": {
                    "force_field": "AMBER",
                    "ph": 7.4,
                    "allow_bad_res": True,
                    "default_altloc": "A"
                }
            },
            "pdb_entries": [],
            "processing": {
                "parallel": False,
                "continue_on_error": True
            }
        }
        
        with open(self.config_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    parts = line.split('|')
                    if len(parts) >= 4:
                        pdb_entry = {
                            "pdb_id": parts[0].strip().upper(),
                            "ligand_selection": {
                                "ligand_name": parts[1].strip() if parts[1].strip() != "AUTO" else None
                            },
                            "cleaning": {
                                "strategy": parts[2].strip()
                            },
                            "autodock": {
                                "prepare_as": parts[3].strip()
                            }
                        }
                        config["pdb_entries"].append(pdb_entry)
                        
        return config
        
    def _get_global_setting(self, key_path: str, default: Any = None) -> Any:
        """
        Get a global setting using dot notation path
        
        Args:
            key_path: Dot-separated path to setting (e.g., "cleaning.strategy")
            default: Default value if not found
            
        Returns:
            Setting value or default
        """
        keys = key_path.split('.')
        value = self.config.get("global_settings", {})
        
        for key in keys:
            if isinstance(value, dict) and key in value:
                value = value[key]
            else:
                return default
                
        return value
        
    def _get_pdb_setting(self, pdb_config: Dict, key_path: str, default: Any = None) -> Any:
        """
        Get a PDB-specific setting, falling back to global setting
        
        Args:
            pdb_config: PDB-specific configuration
            key_path: Dot-separated path to setting
            default: Default value if not found
            
        Returns:
            Setting value or default
        """
        # Check PDB-specific setting first
        keys = key_path.split('.')
        value = pdb_config
        
        for key in keys:
            if isinstance(value, dict) and key in value:
                value = value[key]
            else:
                # Fall back to global setting
                return self._get_global_setting(key_path, default)
                
        return value
        
    def process_pdb_entry(self, pdb_config: Dict) -> bool:
        """
        Process a single PDB entry
        
        Args:
            pdb_config: Configuration for this PDB entry
            
        Returns:
            True if successful, False otherwise
        """
        pdb_id = pdb_config["pdb_id"]
        print(f"\n{'='*60}")
        print(f"Processing PDB: {pdb_id}")
        print(f"{'='*60}")
        
        try:
            # Create PDB-specific output directory
            pdb_output_dir = self.output_dir / pdb_id
            pdb_output_dir.mkdir(exist_ok=True)
            
            # Initialize pipeline
            pipeline = MolecularDockingPipeline(str(pdb_output_dir))
            
            # Step 1: Download PDB
            pdb_file = pipeline.fetch_pdb(pdb_id)
            
            # Step 2: Enumerate HETATMs
            hetatm_details, unique_hetatms = pipeline.enumerate_hetatms(pdb_file)
            
            # Step 3: Select ligand
            selected_hetatm = None
            chain_id = None
            res_id = None
            
            if unique_hetatms:
                ligand_name = self._get_pdb_setting(pdb_config, "ligand_selection.ligand_name")
                
                if ligand_name:
                    # Find specific ligand from config
                    for resname, c_id, r_id, _ in hetatm_details:
                        if resname == ligand_name.upper():
                            selected_hetatm = resname
                            chain_id = c_id
                            res_id = r_id
                            print(f"✓ Using configured ligand: {selected_hetatm}_{chain_id}_{res_id}")
                            break
                else:
                    # Auto-select first available ligand (exclude common ions/water)
                    common_solvents = {'HOH', 'NA', 'CL', 'SO4', 'CA', 'MG', 'ZN', 'FE', 'CU', 'MN'}
                    for resname, c_id, r_id, _ in hetatm_details:
                        if resname not in common_solvents:
                            selected_hetatm = resname
                            chain_id = c_id
                            res_id = r_id
                            print(f"✓ Auto-selected ligand: {selected_hetatm}_{chain_id}_{res_id}")
                            break
                
                if selected_hetatm:
                    # Save ligand as separate PDB
                    ligand_pdb = pipeline.save_hetatm_as_pdb(
                        pdb_file, selected_hetatm, chain_id, res_id
                    )
            
            # Step 4: Clean PDB
            cleaning_strategy = self._get_pdb_setting(pdb_config, "cleaning.strategy")
            
            if isinstance(cleaning_strategy, str):
                if cleaning_strategy == "all":
                    to_remove_list = unique_hetatms
                elif cleaning_strategy == "common":
                    common_residues = self._get_global_setting("cleaning.common_residues", [])
                    to_remove_list = [r for r in common_residues if r in unique_hetatms]
                elif cleaning_strategy == "none":
                    to_remove_list = []
                else:
                    to_remove_list = []
            elif isinstance(cleaning_strategy, list):
                to_remove_list = cleaning_strategy
            else:
                to_remove_list = []
            
            # Remove selected ligand from cleaning list
            if selected_hetatm and selected_hetatm in to_remove_list:
                to_remove_list.remove(selected_hetatm)
                print(f"✓ Removed {selected_hetatm} from cleaning list (selected as ligand)")
            
            cleaned_pdb = pipeline.clean_pdb(pdb_file, to_remove_list)
            
            # Step 5: Extract active site coordinates and analyze
            results = {'pdb_id': pdb_id}
            
            if selected_hetatm:
                try:
                    # Use enhanced residue-level analysis
                    use_enhanced = self._get_pdb_setting(pdb_config, "analysis.use_enhanced_coordinates", True)
                    
                    if use_enhanced:
                        residue_analysis = extract_residue_level_coordinates(
                            cleaned_pdb, selected_hetatm, chain_id, res_id
                        )
                        if residue_analysis:
                            coords = residue_analysis['overall_center']
                            results.update({
                                'selected_ligand': f"{selected_hetatm}_{chain_id}_{res_id}",
                                'active_site_center_x': coords[0],
                                'active_site_center_y': coords[1], 
                                'active_site_center_z': coords[2],
                                'interacting_residues_count': residue_analysis['num_interacting_residues'],
                                'interacting_atoms_count': residue_analysis['num_interacting_atoms']
                            })
                        else:
                            raise ValueError("Enhanced coordinate extraction failed")
                    else:
                        coords, num_atoms = pipeline.extract_active_site_coords(
                            cleaned_pdb, selected_hetatm, chain_id, res_id
                        )
                        results.update({
                            'selected_ligand': f"{selected_hetatm}_{chain_id}_{res_id}",
                            'active_site_center_x': coords[0],
                            'active_site_center_y': coords[1],
                            'active_site_center_z': coords[2],
                            'interacting_atoms_count': num_atoms
                        })
                    
                    # Analyze pocket properties
                    pocket_results = pipeline.analyze_pocket_properties(cleaned_pdb, coords)
                    results.update(pocket_results)
                    
                except Exception as e:
                    print(f"⚠️  Active site analysis failed: {e}")
                    print("Proceeding without pocket analysis.")
            
            # Step 6: Generate reports
            report_file = pipeline.generate_summary_report(results, pdb_id)
            
            # Store results
            self.results[pdb_id] = {
                "success": True,
                "results": results,
                "output_dir": str(pdb_output_dir)
            }
            
            # Step 7: AutoDock preparation if requested
            prepare_as = self._get_pdb_setting(pdb_config, "autodock.prepare_as", "none")
            if prepare_as in ["ligand", "receptor", "both"]:
                self._prepare_for_autodock(pdb_id, pdb_file, pdb_output_dir, pdb_config, prepare_as)
            
            print(f"✅ Successfully processed {pdb_id}")
            return True
            
        except Exception as e:
            print(f"❌ Error processing {pdb_id}: {e}")
            self.results[pdb_id] = {
                "success": False,
                "error": str(e)
            }
            return False
    
    def _prepare_for_autodock(self, pdb_id: str, pdb_file: str, pdb_output_dir: Path, 
                             pdb_config: Dict, prepare_as: str) -> bool:
        """
        Prepare PDB for AutoDock Vina
        
        Args:
            pdb_id: PDB identifier
            pdb_file: Path to PDB file
            pdb_output_dir: Output directory for this PDB
            pdb_config: Configuration for this PDB
            prepare_as: What to prepare as ("ligand", "receptor", "both")
            
        Returns:
            True if successful, False otherwise
        """
        try:
            print(f"🔄 Preparing {pdb_id} for AutoDock...")
            
            # Create AutoDock output directories
            autodock_dir = pdb_output_dir / "autodock_preparation"
            ligands_dir = autodock_dir / "ligands"
            receptors_dir = autodock_dir / "receptors"
            ligands_dir.mkdir(exist_ok=True)
            receptors_dir.mkdir(exist_ok=True)
            
            # Get AutoDock settings
            force_field = self._get_pdb_setting(pdb_config, "autodock.force_field", "AMBER")
            ph = self._get_pdb_setting(pdb_config, "autodock.ph", 7.4)
            allow_bad_res = self._get_pdb_setting(pdb_config, "autodock.allow_bad_res", True)
            default_altloc = self._get_pdb_setting(pdb_config, "autodock.default_altloc", "A")
            
            # Create AutoDock configuration
            config = PreparationConfig(
                ligands_input=str(pdb_output_dir),
                receptors_input=str(pdb_output_dir),
                ligands_output=str(ligands_dir),
                receptors_output=str(receptors_dir),
                force_field=force_field,
                ph=ph,
                allow_bad_res=allow_bad_res,
                default_altloc=default_altloc
            )
            
            # Initialize AutoDock pipeline
            autodock_pipeline = AutoDockPreparationPipeline(config)
            
            # Prepare based on request
            if prepare_as in ["ligand", "both"]:
                # For ligand preparation, we need to extract the ligand first
                # This is already done in the main pipeline, so we can use that file
                pass  # The ligand PDB file is already created
                
            if prepare_as in ["receptor", "both"]:
                # Prepare receptor
                # Copy the cleaned PDB to receptors input directory
                import shutil
                receptor_input = receptors_dir / f"{pdb_id}.pdb"
                shutil.copy2(pdb_file, receptor_input)
            
            # For now, we'll just note that AutoDock preparation was requested
            # A full implementation would integrate with the existing AutoDock preparation
            print(f"✅ AutoDock preparation requested for {pdb_id} as {prepare_as}")
            print(f"   Output directory: {autodock_dir}")
            return True
            
        except Exception as e:
            print(f"❌ AutoDock preparation failed for {pdb_id}: {e}")
            return False
    
    def run_batch_processing(self) -> Dict[str, Any]:
        """
        Run batch processing for all PDB entries
        
        Returns:
            Dictionary with processing results
        """
        print(f"🚀 Starting Batch PDB Processing")
        print(f"Configuration file: {self.config_file}")
        print(f"Output directory: {self.output_dir}")
        
        pdb_entries = self.config.get("pdb_entries", [])
        if not pdb_entries:
            print("❌ No PDB entries found in configuration")
            return {"success": False, "error": "No PDB entries found"}
            
        print(f"Found {len(pdb_entries)} PDB entries to process")
        
        # Process PDB entries
        successful = 0
        failed = 0
        
        continue_on_error = self.config.get("processing", {}).get("continue_on_error", True)
        
        for i, pdb_config in enumerate(pdb_entries, 1):
            pdb_id = pdb_config.get("pdb_id")
            if not pdb_id:
                print(f"⚠️  Skipping entry {i} - no PDB ID specified")
                failed += 1
                continue
                
            print(f"\nProcessing PDB {i}/{len(pdb_entries)}: {pdb_id}")
            
            try:
                success = self.process_pdb_entry(pdb_config)
                if success:
                    successful += 1
                else:
                    failed += 1
                    
                if not success and not continue_on_error:
                    print(f"❌ Stopping batch processing due to error in {pdb_id}")
                    break
                    
            except Exception as e:
                print(f"❌ Unexpected error processing {pdb_id}: {e}")
                failed += 1
                if not continue_on_error:
                    break
        
        # Generate batch summary report
        self._generate_batch_report()
        
        # Summary
        print(f"\n📊 Batch Processing Summary:")
        print(f"   Total PDBs: {len(pdb_entries)}")
        print(f"   Successful: {successful}")
        print(f"   Failed: {failed}")
        
        if failed > 0:
            print(f"\n❌ Failed PDBs:")
            for pdb_id, result in self.results.items():
                if not result.get("success", False):
                    print(f"   - {pdb_id}: {result.get('error', 'Unknown error')}")
        
        print(f"\n✓ Batch processing completed!")
        return {
            "success": True,
            "total": len(pdb_entries),
            "successful": successful,
            "failed": failed,
            "results": self.results
        }
    
    def _generate_batch_report(self):
        """
        Generate a comprehensive batch processing report
        """
        try:
            # Create Excel report if possible
            if EXCEL_AVAILABLE:
                excel_path = self.output_dir / "batch_processing_results.xlsx"
                wb = Workbook()
                
                # Remove default sheet
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                
                # Create summary sheet
                summary_ws = wb.create_sheet("Summary")
                summary_headers = ["PDB ID", "Status", "Selected Ligand", "Interacting Residues", 
                                 "Pocket Volume (Å³)", "Druggability Score", "Output Directory"]
                summary_ws.append(summary_headers)
                
                # Style headers
                for cell in summary_ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data
                for pdb_id, result in self.results.items():
                    if result.get("success", False):
                        results_data = result.get("results", {})
                        row = [
                            pdb_id,
                            "Success",
                            results_data.get("selected_ligand", "N/A"),
                            results_data.get("interacting_residues_count", "N/A"),
                            results_data.get("pocket_volume_A3", "N/A"),
                            results_data.get("druggability_score", "N/A"),
                            result.get("output_dir", "N/A")
                        ]
                    else:
                        row = [
                            pdb_id,
                            "Failed",
                            "N/A",
                            "N/A",
                            "N/A",
                            "N/A",
                            result.get("error", "Unknown error")
                        ]
                    summary_ws.append(row)
                
                # Save workbook
                wb.save(excel_path)
                print(f"📊 Excel report saved: {excel_path}")
            
            # Create CSV summary
            csv_data = []
            for pdb_id, result in self.results.items():
                if result.get("success", False):
                    results_data = result.get("results", {})
                    csv_data.append({
                        "PDB_ID": pdb_id,
                        "Status": "Success",
                        "Selected_Ligand": results_data.get("selected_ligand", "N/A"),
                        "Interacting_Residues": results_data.get("interacting_residues_count", "N/A"),
                        "Pocket_Volume_A3": results_data.get("pocket_volume_A3", "N/A"),
                        "Druggability_Score": results_data.get("druggability_score", "N/A"),
                        "Output_Directory": result.get("output_dir", "N/A")
                    })
                else:
                    csv_data.append({
                        "PDB_ID": pdb_id,
                        "Status": "Failed",
                        "Selected_Ligand": "N/A",
                        "Interacting_Residues": "N/A",
                        "Pocket_Volume_A3": "N/A",
                        "Druggability_Score": "N/A",
                        "Output_Directory": result.get("error", "Unknown error")
                    })
            
            csv_path = self.output_dir / "batch_processing_results.csv"
            df = pd.DataFrame(csv_data)
            df.to_csv(csv_path, index=False)
            print(f"📋 CSV report saved: {csv_path}")
            
        except Exception as e:
            print(f"⚠️  Failed to generate batch report: {e}")


def main():
    """
    Main function for command-line usage
    """
    parser = argparse.ArgumentParser(
        description="Batch PDB Preparation Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process PDBs using YAML configuration
  python batch_pdb_preparation.py -c pdb_batch_config.yaml
  
  # Process PDBs using TXT configuration
  python batch_pdb_preparation.py -c pdb_batch_config.txt
  
  # Specify output directory
  python batch_pdb_preparation.py -c pdb_batch_config.yaml -o batch_results
        """
    )
    
    parser.add_argument('-c', '--config', type=str, required=True,
                       help='Configuration file (YAML or TXT format)')
    parser.add_argument('-o', '--output', type=str, default='batch_docking_preparation',
                       help='Output directory (default: batch_docking_preparation)')
    parser.add_argument('-v', '--verbose', action='store_true',
                       help='Enable verbose output')
    
    args = parser.parse_args()
    
    try:
        # Initialize batch pipeline
        batch_pipeline = BatchPDBPreparationPipeline(args.config, args.output)
        
        # Run batch processing
        results = batch_pipeline.run_batch_processing()
        
        # Exit with appropriate code
        sys.exit(0 if results.get("successful", 0) > 0 else 1)
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Batch processing interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()