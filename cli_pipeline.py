#!/usr/bin/env python3
"""
PDB Prepare Wizard - Command Line Interface
==========================================

Command-line interface for the PDB preparation pipeline with batch processing capabilities.

Author: Molecular Docking Pipeline
Version: 2.1.0
"""

import argparse
import sys
import json
from pathlib import Path
from typing import List, Optional, Dict, Any

# Import the core pipeline
from core_pipeline import MolecularDockingPipeline, extract_residue_level_coordinates, EXCEL_AVAILABLE

if EXCEL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

def load_config(config_file: str) -> Dict[str, Any]:
    """Load configuration from JSON file."""
    try:
        with open(config_file, 'r') as f:
            config = json.load(f)
        print(f"✓ Configuration loaded from: {config_file}")
        return config
    except Exception as e:
        print(f"❌ Failed to load configuration: {e}")
        return {}

def parse_pdb_list(pdb_input: str) -> List[str]:
    """Parse PDB list from various input formats."""
    if ',' in pdb_input:
        # Comma-separated list
        return [pdb.strip().upper() for pdb in pdb_input.split(',') if pdb.strip()]
    elif Path(pdb_input).exists():
        # File with PDB IDs
        try:
            with open(pdb_input, 'r') as f:
                return [line.strip().upper() for line in f if line.strip() and not line.startswith('#')]
        except Exception as e:
            print(f"❌ Failed to read PDB file: {e}")
            return []
    else:
        # Single PDB ID
        return [pdb_input.strip().upper()]

def validate_pdb_id(pdb_id: str) -> bool:
    """Validate PDB ID format."""
    return len(pdb_id) == 4 and pdb_id.isalnum()

def run_batch_analysis(pipeline: MolecularDockingPipeline, pdb_list: List[str], 
                      config: Dict[str, Any], output_dir: Path) -> Dict[str, bool]:
    """Run batch analysis for multiple PDBs."""
    results = {}
    excel_workbook = None
    
    if len(pdb_list) > 1 and EXCEL_AVAILABLE:
        excel_workbook = Workbook()
        if "Sheet" in excel_workbook.sheetnames:
            excel_workbook.remove(excel_workbook["Sheet"])
        print("✓ Excel workbook created for batch analysis")
    
    for i, pdb_id in enumerate(pdb_list, 1):
        print(f"\n{'='*60}")
        print(f"Processing PDB {i}/{len(pdb_list)}: {pdb_id}")
        print(f"{'='*60}")
        
        if not validate_pdb_id(pdb_id):
            print(f"❌ Invalid PDB ID format: {pdb_id}")
            results[pdb_id] = False
            continue
        
        try:
            # Run single PDB analysis
            success = run_single_pdb_cli(pipeline, pdb_id, config, excel_workbook)
            results[pdb_id] = success
            
            if success:
                print(f"✅ Successfully processed {pdb_id}")
            else:
                print(f"❌ Failed to process {pdb_id}")
                
        except Exception as e:
            print(f"❌ Error processing {pdb_id}: {e}")
            results[pdb_id] = False
    
    # Save Excel file for batch analysis
    if excel_workbook and EXCEL_AVAILABLE and len(pdb_list) > 1:
        excel_path = output_dir / "batch_analysis_results.xlsx"
        excel_workbook.save(excel_path)
        print(f"\n📊 Batch analysis results saved to: {excel_path}")
    
    return results

def run_single_pdb_cli(pipeline: MolecularDockingPipeline, pdb_id: str, 
                      config: Dict[str, Any], excel_workbook=None) -> bool:
    """Run analysis for a single PDB using CLI configuration."""
    try:
        # Step 1: Download PDB
        pdb_file = pipeline.fetch_pdb(pdb_id)
        
        # Step 2: Enumerate HETATMs
        hetatm_details, unique_hetatms = pipeline.enumerate_hetatms(pdb_file)
        
        # Step 3: Select ligand (from config or first available)
        selected_hetatm = None
        chain_id = None
        res_id = None
        
        if unique_hetatms:
            ligand_config = config.get('ligand_selection', {})
            target_ligand = ligand_config.get(pdb_id.lower())
            
            if target_ligand:
                # Find specific ligand from config
                for resname, c_id, r_id, _ in hetatm_details:
                    if resname == target_ligand.upper():
                        selected_hetatm = resname
                        chain_id = c_id
                        res_id = r_id
                        print(f"✓ Using configured ligand: {selected_hetatm}_{chain_id}_{res_id}")
                        break
            
            if not selected_hetatm:
                # Use first available ligand (exclude common ions/water)
                common_solvents = {'HOH', 'NA', 'CL', 'SO4', 'CA', 'MG', 'ZN', 'FE', 'CU', 'MN'}
                for resname, c_id, r_id, _ in hetatm_details:
                    if resname not in common_solvents:
                        selected_hetatm = resname
                        chain_id = c_id
                        res_id = r_id
                        print(f"✓ Auto-selected ligand: {selected_hetatm}_{chain_id}_{res_id}")
                        break
            
            if selected_hetatm:
                # Save ligand as separate PDB
                ligand_pdb = pipeline.save_hetatm_as_pdb(
                    pdb_file, selected_hetatm, chain_id, res_id
                )
        
        # Step 4: Clean PDB
        cleaning_config = config.get('cleaning', {})
        pdb_specific_cleaning = cleaning_config.get(pdb_id.lower())
        
        if pdb_specific_cleaning:
            to_remove_list = pdb_specific_cleaning
        else:
            # Default cleaning strategy
            default_strategy = cleaning_config.get('default', 'common')
            if default_strategy == 'all':
                to_remove_list = unique_hetatms
            elif default_strategy == 'common':
                common_residues = ['HOH', 'NA', 'CL', 'SO4', 'CA', 'MG', 'ZN', 'FE', 'CU', 'MN']
                to_remove_list = [r for r in common_residues if r in unique_hetatms]
            elif default_strategy == 'none':
                to_remove_list = []
            else:
                to_remove_list = default_strategy if isinstance(default_strategy, list) else []
        
        # Remove selected ligand from cleaning list
        if selected_hetatm and selected_hetatm in to_remove_list:
            to_remove_list.remove(selected_hetatm)
            print(f"✓ Removed {selected_hetatm} from cleaning list (selected as ligand)")
        
        cleaned_pdb = pipeline.clean_pdb(pdb_file, to_remove_list)
        
        # Step 5: Extract active site coordinates and analyze
        results = {'pdb_id': pdb_id}
        
        if selected_hetatm:
            try:
                # Use enhanced residue-level analysis
                use_enhanced = config.get('analysis', {}).get('use_enhanced_coordinates', True)
                
                if use_enhanced:
                    residue_analysis = extract_residue_level_coordinates(
                        cleaned_pdb, selected_hetatm, chain_id, res_id
                    )
                    if residue_analysis:
                        coords = residue_analysis['overall_center']
                        results.update({
                            'selected_ligand': f"{selected_hetatm}_{chain_id}_{res_id}",
                            'active_site_center_x': coords[0],
                            'active_site_center_y': coords[1], 
                            'active_site_center_z': coords[2],
                            'interacting_residues_count': residue_analysis['num_interacting_residues'],
                            'interacting_atoms_count': residue_analysis['num_interacting_atoms']
                        })
                    else:
                        raise ValueError("Enhanced coordinate extraction failed")
                else:
                    coords, num_atoms = pipeline.extract_active_site_coords(
                        cleaned_pdb, selected_hetatm, chain_id, res_id
                    )
                    results.update({
                        'selected_ligand': f"{selected_hetatm}_{chain_id}_{res_id}",
                        'active_site_center_x': coords[0],
                        'active_site_center_y': coords[1],
                        'active_site_center_z': coords[2],
                        'interacting_atoms_count': num_atoms
                    })
                
                # Analyze pocket properties
                pocket_results = pipeline.analyze_pocket_properties(cleaned_pdb, coords)
                results.update(pocket_results)
                
            except Exception as e:
                print(f"⚠️  Active site analysis failed: {e}")
                print("Proceeding without pocket analysis.")
        
        # Step 6: Generate reports
        report_file = pipeline.generate_summary_report(results, pdb_id)
        
        # Add to Excel workbook if provided
        if excel_workbook and EXCEL_AVAILABLE:
            add_to_excel_workbook(excel_workbook, pdb_id, results)
        
        return True
        
    except Exception as e:
        print(f"❌ Analysis failed for {pdb_id}: {e}")
        return False

def add_to_excel_workbook(workbook, pdb_id, results):
    """Add results to Excel workbook."""
    try:
        # Create or get worksheet
        if "Summary" not in workbook.sheetnames:
            ws = workbook.create_sheet("Summary")
            # Add headers
            headers = ["PDB_ID", "Property", "Value"]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        else:
            ws = workbook["Summary"]
        
        # Add results
        for key, value in results.items():
            ws.append([pdb_id, key, value])
            
        # Add separator row
        ws.append(["", "", ""])
        
    except Exception as e:
        print(f"⚠️  Failed to add results to Excel: {e}")

def create_sample_config():
    """Create a sample configuration file."""
    sample_config = {
        "description": "Sample configuration for PDB Prepare Wizard CLI",
        "ligand_selection": {
            "7cmd": "TTT",
            "6wx4": "LIG",
            "1abc": "GDP"
        },
        "cleaning": {
            "default": "common",
            "7cmd": ["HOH", "NA", "CL"],
            "custom_pdb": "all"
        },
        "analysis": {
            "use_enhanced_coordinates": True,
            "distance_cutoff": 5.0
        }
    }
    
    config_path = "sample_config.json"
    with open(config_path, 'w') as f:
        json.dump(sample_config, f, indent=2)
    
    print(f"✓ Sample configuration created: {config_path}")
    return config_path

def main():
    """Main CLI function."""
    parser = argparse.ArgumentParser(
        description="PDB Prepare Wizard - Command Line Interface",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Analyze single PDB
  python cli_pipeline.py -p 7CMD
  
  # Analyze multiple PDBs
  python cli_pipeline.py -p "7CMD,6WX4,1ABC"
  
  # Use configuration file
  python cli_pipeline.py -p 7CMD -c config.json
  
  # Read PDB list from file
  python cli_pipeline.py -p pdb_list.txt
  
  # Generate sample config
  python cli_pipeline.py --create-config
        """
    )
    
    parser.add_argument('-p', '--pdbs', type=str,
                       help='PDB ID(s) - single ID, comma-separated list, or file path')
    parser.add_argument('-o', '--output', type=str, default='pipeline_output',
                       help='Output directory (default: pipeline_output)')
    parser.add_argument('-c', '--config', type=str,
                       help='Configuration file (JSON format)')
    parser.add_argument('--create-config', action='store_true',
                       help='Create a sample configuration file')
    parser.add_argument('-v', '--verbose', action='store_true',
                       help='Enable verbose output')
    
    args = parser.parse_args()
    
    # Create sample config if requested
    if args.create_config:
        create_sample_config()
        return
    
    # Validate arguments
    if not args.pdbs:
        parser.error("PDB ID(s) are required. Use -p/--pdbs or --create-config")
    
    # Load configuration
    config = {}
    if args.config:
        config = load_config(args.config)
    
    # Parse PDB list
    pdb_list = parse_pdb_list(args.pdbs)
    if not pdb_list:
        print("❌ No valid PDB IDs provided")
        sys.exit(1)
    
    print(f"✓ Found {len(pdb_list)} PDB(s) to process: {', '.join(pdb_list)}")
    
    # Initialize pipeline
    pipeline = MolecularDockingPipeline(args.output)
    output_dir = Path(args.output)
    
    # Run analysis
    print(f"\n🚀 Starting PDB Prepare Wizard CLI")
    print(f"Output directory: {output_dir.absolute()}")
    print(f"PDBs to process: {len(pdb_list)}")
    
    results = run_batch_analysis(pipeline, pdb_list, config, output_dir)
    
    # Summary
    successful = sum(1 for success in results.values() if success)
    failed = len(results) - successful
    
    print(f"\n📊 Analysis Summary:")
    print(f"   Total PDBs: {len(results)}")
    print(f"   Successful: {successful}")
    print(f"   Failed: {failed}")
    
    if failed > 0:
        print(f"\n❌ Failed PDBs:")
        for pdb_id, success in results.items():
            if not success:
                print(f"   - {pdb_id}")
    
    print(f"\n✓ CLI analysis completed!")
    
    # Exit with appropriate code
    sys.exit(0 if failed == 0 else 1)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Analysis interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        sys.exit(1)

