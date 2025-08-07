#!/usr/bin/env python3
"""
PDB Prepare Wizard - Molecular Docking Pipeline
===============================================

A comprehensive tool for preparing PDB files for molecular docking studies.
This script provides functionality to:
- Download PDB files from RCSB PDB database
- Extract and analyze ligands (HETATMs)
- Clean PDB files by removing unwanted residues
- Analyze pocket properties and druggability
- Generate comprehensive reports

Author: Molecular Docking Pipeline
Version: 2.0
"""

import os
import sys
import subprocess
import numpy as np
import pandas as pd
from pathlib import Path
from typing import List, Tuple, Dict, Optional, Any
import warnings
warnings.filterwarnings("ignore")

# Add Excel support
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not available - Excel output will be disabled")

try:
    from Bio.PDB import PDBList, PDBParser, Select, PDBIO
    from Bio.PDB.Structure import Structure
    from Bio.PDB.Model import Model
    from Bio.PDB.Chain import Chain
except ImportError as e:
    print(f"❌ Error importing Biopython: {e}")
    print("Please install Biopython: pip install biopython")
    sys.exit(1)

class MolecularDockingPipeline:
    """
    A comprehensive molecular docking pipeline for PDB file preparation and analysis.
    
    This class provides methods to:
    - Download and process PDB files
    - Extract ligands and analyze binding sites
    - Clean and prepare structures for docking
    - Analyze pocket properties and druggability
    """
    
    def __init__(self, output_dir: str = "pipeline_output"):
        """
        Initialize the molecular docking pipeline.
        
        Args:
            output_dir (str): Directory to store all output files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        print(f"✓ Pipeline initialized. Output directory: {self.output_dir}")
        
        # Validate required packages
        self._validate_dependencies()
        
    def _validate_dependencies(self):
        """Validate that all required dependencies are available."""
        try:
            import numpy as np
            import pandas as pd
            from Bio.PDB import PDBParser
            print("✓ All core dependencies available")
        except ImportError as e:
            print(f"❌ Missing dependency: {e}")
            sys.exit(1)
            
        # Check for optional PLIP
        try:
            import plip
            print("✓ PLIP available for advanced interaction analysis")
            self.plip_available = True
        except ImportError:
            print("⚠️  PLIP not available - will use distance-based analysis")
            self.plip_available = False
    
    def fetch_pdb(self, pdb_id: str) -> str:
        """
        Download PDB file from RCSB PDB database.
        
        Args:
            pdb_id (str): PDB identifier (e.g., '1ABC')
            
        Returns:
            str: Path to the downloaded PDB file
            
        Raises:
            ValueError: If PDB ID is invalid or download fails
        """
        if not pdb_id or len(pdb_id) != 4:
            raise ValueError("PDB ID must be exactly 4 characters")
            
        print(f"🔄 Fetching PDB {pdb_id}...")
        
        try:
            pdbl = PDBList()
            filename = pdbl.retrieve_pdb_file(
                pdb_id.lower(), 
                pdir=str(self.output_dir), 
                file_format='pdb'
            )
            
            # Rename to simpler format
            new_filename = self.output_dir / f"{pdb_id.upper()}.pdb"
            os.rename(filename, new_filename)
            
            if not new_filename.exists():
                raise FileNotFoundError(f"Failed to download PDB {pdb_id}")
                
            print(f"✓ Downloaded: {new_filename}")
            return str(new_filename)
            
        except Exception as e:
            raise ValueError(f"Failed to download PDB {pdb_id}: {e}")
    
    def enumerate_hetatms(self, pdb_file: str) -> Tuple[List[Tuple], List[str]]:
        """
        List all HETATM residues in the PDB file for user selection.
        
        Args:
            pdb_file (str): Path to PDB file
            
        Returns:
            Tuple[List[Tuple], List[str]]: 
                - List of (resname, chain_id, res_id, residue) tuples
                - List of unique HETATM residue names
        """
        print(f"🔄 Enumerating HETATMs in {pdb_file}...")
        
        if not Path(pdb_file).exists():
            raise FileNotFoundError(f"PDB file not found: {pdb_file}")
        
        try:
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure('protein', pdb_file)
            
            hetatms = []
            hetatm_details = []
            hetatm_counts = {}
            
            for model in structure:
                for chain in model:
                    for residue in chain:
                        if residue.get_id()[0] != ' ':  # HETATM check
                            resname = residue.get_resname()
                            chain_id = chain.get_id()
                            res_id = residue.get_id()[1]
                            detail = f"{resname}_{chain_id}_{res_id}"
                            hetatms.append(resname)
                            hetatm_details.append((resname, chain_id, res_id, residue))
                            
                            # Count occurrences
                            hetatm_counts[resname] = hetatm_counts.get(resname, 0) + 1
            
            unique_hetatms = list(set(hetatms))
            
            # Group and display HETATMs by type
            print(f"✓ Found HETATM types: {len(unique_hetatms)}")
            print(f"✓ Total HETATM instances: {len(hetatm_details)}")
            print("\n📋 HETATM Summary:")
            for resname in sorted(unique_hetatms):
                count = hetatm_counts[resname]
                print(f"   - {resname}: {count} instance(s)")
            
            return hetatm_details, unique_hetatms
            
        except Exception as e:
            raise ValueError(f"Failed to enumerate HETATMs: {e}")
    
    def save_hetatm_as_pdb(self, pdb_file: str, selected_hetatm: str, 
                           chain_id: str, res_id: int, 
                           output_filename: Optional[str] = None) -> Optional[str]:
        """
        Save selected HETATM as separate PDB file.
        
        Args:
            pdb_file (str): Path to source PDB file
            selected_hetatm (str): HETATM residue name
            chain_id (str): Chain identifier
            res_id (int): Residue ID
            output_filename (str, optional): Output filename
            
        Returns:
            Optional[str]: Path to saved ligand PDB file, or None if failed
        """
        print(f"🔄 Saving HETATM {selected_hetatm}_{chain_id}_{res_id} as separate PDB...")
        
        try:
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure('protein', pdb_file)
            
            # Create new structure containing only the selected HETATM
            new_structure = Structure('ligand')
            new_model = Model(0)
            new_chain = Chain(chain_id)
            
            # Find and copy the specific HETATM residue
            found_residue = None
            for model in structure:
                for chain in model:
                    if chain.get_id() == chain_id:
                        for residue in chain:
                            if (residue.get_resname() == selected_hetatm and 
                                residue.get_id()[1] == res_id and
                                residue.get_id()[0] != ' '):
                                found_residue = residue
                                break
            
            if found_residue:
                new_chain.add(found_residue.copy())
                new_model.add(new_chain)
                new_structure.add(new_model)
                
                # Save as PDB
                if output_filename is None:
                    output_filename = self.output_dir / f"ligand_{selected_hetatm}_{chain_id}_{res_id}.pdb"
                
                io = PDBIO()
                io.set_structure(new_structure)
                io.save(str(output_filename))
                print(f"✓ Ligand saved as: {output_filename}")
                return str(output_filename)
            else:
                print(f"❌ Could not find HETATM {selected_hetatm}_{chain_id}_{res_id}")
                return None
                
        except Exception as e:
            print(f"❌ Error saving HETATM: {e}")
            return None

    def manage_chains(self, pdb_file: str, output_filename: Optional[str] = None) -> str:
        """
        Allow user to select which chains to keep or remove.
        
        Args:
            pdb_file (str): Path to PDB file
            output_filename (str, optional): Output filename
            
        Returns:
            str: Path to filtered PDB file
        """
        print("🔄 Chain Management...")
        
        try:
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure('protein', pdb_file)
            
            # List all available chains
            all_chains = []
            for model in structure:
                for chain in model:
                    chain_id = chain.get_id()
                    num_residues = len([r for r in chain if r.get_id()[0] == ' '])
                    num_hetatms = len([r for r in chain if r.get_id()[0] != ' '])
                    all_chains.append((chain_id, num_residues, num_hetatms))
            
            print("📋 Available chains:")
            for chain_id, num_res, num_het in all_chains:
                print(f"   Chain {chain_id}: {num_res} protein residues, {num_het} HETATMs")
            
            keep_chains = input("Enter chains to KEEP (comma-separated, e.g., A,B): ").strip().upper().split(',')
            keep_chains = [c.strip() for c in keep_chains if c.strip()]
            
            print(f"✓ Keeping chains: {keep_chains}")
            
            # Create new structure with only selected chains
            new_structure = Structure('filtered')
            new_model = Model(0)
            
            for model in structure:
                for chain in model:
                    if chain.get_id() in keep_chains:
                        new_model.add(chain.copy())
            
            new_structure.add(new_model)
            
            if output_filename is None:
                output_filename = self.output_dir / "chain_filtered.pdb"
            
            io = PDBIO()
            io.set_structure(new_structure)
            io.save(str(output_filename))
            
            print(f"✓ Filtered PDB saved as: {output_filename}")
            return str(output_filename)
            
        except Exception as e:
            raise ValueError(f"Chain management failed: {e}")

    def clean_pdb(self, pdb_file: str, to_remove_list: Optional[List[str]] = None, 
                  output_filename: Optional[str] = None) -> str:
        """
        Clean PDB by removing specified residues (e.g., water, ions).
        
        Args:
            pdb_file (str): Path to PDB file
            to_remove_list (List[str], optional): List of residues to remove
            output_filename (str, optional): Output filename
            
        Returns:
            str: Path to cleaned PDB file
        """
        if to_remove_list is None:
            # First, enumerate all HETATMs in the structure
            hetatm_details, unique_hetatms = self.enumerate_hetatms(pdb_file)
            
            print("📝 Cleaning Options:")
            print("   1. Remove specific residues (enter comma-separated list)")
            print("   2. Remove ALL HETATMs (enter 'ALL')")
            print("   3. Remove common residues only (enter 'COMMON')")
            print("   4. Remove water only (enter 'WATER')")
            print("   5. Remove ions only (enter 'IONS')")
            
            choice = input("Enter your choice (1-5, or comma-separated list): ").strip().upper()
            
            if choice == 'ALL':
                # Remove all HETATMs
                to_remove_list = unique_hetatms
                print(f"🔄 Removing ALL HETATMs: {unique_hetatms}")
            elif choice == 'COMMON':
                # Remove common residues
                common_residues = ['HOH', 'NA', 'CL', 'SO4', 'CA', 'MG', 'ZN', 'FE', 'CU', 'MN']
                to_remove_list = [r for r in common_residues if r in unique_hetatms]
                print(f"🔄 Removing common residues: {to_remove_list}")
            elif choice == 'WATER':
                # Remove water only
                water_residues = ['HOH', 'WAT', 'TIP3', 'SPC']
                to_remove_list = [r for r in water_residues if r in unique_hetatms]
                print(f"🔄 Removing water molecules: {to_remove_list}")
            elif choice == 'IONS':
                # Remove ions only
                ion_residues = ['NA', 'CL', 'CA', 'MG', 'ZN', 'FE', 'CU', 'MN', 'K', 'LI']
                to_remove_list = [r for r in ion_residues if r in unique_hetatms]
                print(f"🔄 Removing ions: {to_remove_list}")
            else:
                # Parse comma-separated list
                to_remove_list = [r.strip().upper() for r in choice.split(',') if r.strip()]
                print(f"🔄 Removing specified residues: {to_remove_list}")
            
            # Show summary of what will be removed
            self._show_removal_summary(pdb_file, to_remove_list)
            
            # Check if selected ligand will be removed
            if hasattr(self, 'selected_ligand_info'):
                selected_resname = self.selected_ligand_info.get('resname')
                if selected_resname and selected_resname in to_remove_list:
                    print(f"⚠️  WARNING: Your selected ligand ({selected_resname}) will be removed!")
                    print("   This may affect the active site analysis.")
                    warning_confirm = input("Continue anyway? (y/n): ").lower().strip()
                    if warning_confirm != 'y':
                        print("❌ Cleaning cancelled by user")
                        return pdb_file
            
            # Ask for confirmation
            confirm = input("Proceed with removal? (y/n): ").lower().strip()
            if confirm != 'y':
                print("❌ Cleaning cancelled by user")
                return pdb_file
        
        print(f"🔄 Cleaning PDB - removing: {to_remove_list}")
        
        try:
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure('protein', pdb_file)
            
            if output_filename is None:
                output_filename = self.output_dir / "cleaned.pdb"
            
            class RemoveSelect(Select):
                def __init__(self, to_remove):
                    self.to_remove = to_remove
                
                def accept_residue(self, residue):
                    return residue.get_resname() not in self.to_remove
            
            io = PDBIO()
            io.set_structure(structure)
            io.save(str(output_filename), RemoveSelect(to_remove_list))
            
            print(f"✓ Cleaned PDB saved as: {output_filename}")
            return str(output_filename)
            
        except Exception as e:
            raise ValueError(f"PDB cleaning failed: {e}")
    
    def _show_removal_summary(self, pdb_file: str, to_remove_list: List[str]):
        """
        Show summary of residues that will be removed.
        
        Args:
            pdb_file (str): Path to PDB file
            to_remove_list (List[str]): List of residues to remove
        """
        try:
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure('protein', pdb_file)
            
            removal_counts = {}
            total_removals = 0
            
            for model in structure:
                for chain in model:
                    for residue in chain:
                        if residue.get_resname() in to_remove_list:
                            resname = residue.get_resname()
                            removal_counts[resname] = removal_counts.get(resname, 0) + 1
                            total_removals += 1
            
            print(f"\n📊 Removal Summary:")
            print(f"   Total residues to remove: {total_removals}")
            for resname, count in removal_counts.items():
                print(f"   - {resname}: {count} residues")
            
        except Exception as e:
            print(f"⚠️  Could not generate removal summary: {e}")

    def distance_based_interaction_detection(self, pdb_file: str, ligand_name: str, 
                                          chain_id: str, res_id: int, 
                                          cutoff: float = 5.0) -> List[np.ndarray]:
        """
        Detect interactions using distance-based method when PLIP is not available.
        
        Args:
            pdb_file (str): Path to PDB file
            ligand_name (str): Ligand residue name
            chain_id (str): Chain identifier
            res_id (int): Residue ID
            cutoff (float): Distance cutoff in Angstroms
            
        Returns:
            List[np.ndarray]: List of coordinates of interacting atoms
        """
        print(f"🔄 Using distance-based interaction detection (cutoff: {cutoff}Å)...")
        
        try:
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure('protein', pdb_file)
            
            # Find the ligand
            ligand_residue = None
            for model in structure:
                for chain in model:
                    if chain.get_id() == chain_id:
                        for residue in chain:
                            if (residue.get_resname() == ligand_name and 
                                residue.get_id()[1] == res_id and
                                residue.get_id()[0] != ' '):
                                ligand_residue = residue
                                break
            
            if not ligand_residue:
                raise ValueError(f"Ligand {ligand_name}_{chain_id}_{res_id} not found")
            
            # Get ligand atom coordinates
            ligand_coords = []
            for atom in ligand_residue:
                ligand_coords.append(atom.get_coord())
            
            # Find interacting protein residues
            interacting_residues = []
            all_coords = []
            
            for model in structure:
                for chain in model:
                    for residue in chain:
                        if residue.get_id()[0] == ' ':  # Protein residue
                            for atom in residue:
                                atom_coord = atom.get_coord()
                                # Check distance to any ligand atom
                                for lig_coord in ligand_coords:
                                    dist = np.linalg.norm(atom_coord - lig_coord)
                                    if dist <= cutoff:
                                        if residue not in interacting_residues:
                                            interacting_residues.append(residue)
                                        all_coords.append(atom_coord)
                                        break
            
            print(f"✓ Found {len(interacting_residues)} interacting residues")
            return all_coords
            
        except Exception as e:
            raise ValueError(f"Distance-based interaction detection failed: {e}")

    def extract_active_site_coords(self, cleaned_pdb: str, ligand_name: str, 
                                 chain_id: str, res_id: int, 
                                 method: str = 'distance') -> Tuple[np.ndarray, int]:
        """
        Extract active site coordinates using specified method.
        
        Args:
            cleaned_pdb (str): Path to cleaned PDB file
            ligand_name (str): Ligand residue name
            chain_id (str): Chain identifier
            res_id (int): Residue ID
            method (str): Method to use ('plip' or 'distance')
            
        Returns:
            Tuple[np.ndarray, int]: Active site center coordinates and number of atoms
        """
        print(f"🔄 Extracting active site coordinates using {method} method...")
        
        coords = []
        
        if method == 'plip' and self.plip_available:
            try:
                # Try using PLIP first
                from plip.structure.preparation import PDBComplex
                
                my_mol = PDBComplex()
                my_mol.load_pdb(cleaned_pdb)
                my_mol.analyze()
                
                interactions = my_mol.interaction_sets
                if not interactions:
                    print("⚠️  No interactions found with PLIP, falling back to distance method")
                    method = 'distance'
                else:
                    for site in interactions.values():
                        # Get all atoms from interacting residues
                        for res_atoms in site.nearby_residues:
                            for atom in res_atoms:
                                coords.append(atom.get_coord())
                    print(f"✓ PLIP found {len(coords)} interacting atoms")
            except Exception as e:
                print(f"⚠️  PLIP failed: {e}, using distance method")
                method = 'distance'
        
        if method == 'distance' or len(coords) == 0:
            coords = self.distance_based_interaction_detection(
                cleaned_pdb, ligand_name, chain_id, res_id
            )
        
        if not coords:
            raise ValueError("No coordinates extracted")
        
        # Calculate average XYZ
        avg_xyz = np.mean(coords, axis=0)
        print(f"✓ Active site center: X={avg_xyz[0]:.2f}, Y={avg_xyz[1]:.2f}, Z={avg_xyz[2]:.2f}")
        
        return avg_xyz, len(coords)

    def analyze_pocket_properties(self, cleaned_pdb: str, center_coords: np.ndarray, 
                                ligand_pdb: Optional[str] = None) -> Dict[str, Any]:
        """
        Comprehensive pocket analysis using multiple approaches.
        
        Args:
            cleaned_pdb (str): Path to cleaned PDB file
            center_coords (np.ndarray): Pocket center coordinates
            ligand_pdb (str, optional): Path to ligand PDB file
            
        Returns:
            Dict[str, Any]: Dictionary containing pocket analysis results
        """
        print("🔄 Starting comprehensive pocket analysis...")
        
        results = {
            'center_x': center_coords[0],
            'center_y': center_coords[1], 
            'center_z': center_coords[2]
        }
        
        try:
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure('protein', cleaned_pdb)
            
            # 1. Pocket Size and Shape Analysis
            print("📊 Analyzing pocket size and shape...")
            try:
                # Geometric estimation based on interaction sphere
                results['pocket_volume_A3'] = 4/3 * np.pi * (5.0**3)  # Assume 5Å interaction sphere
            except Exception as e:
                print(f"⚠️  Pocket volume analysis failed: {e}")
                results['pocket_volume_A3'] = 'N/A'
            
            # 2. Electrostatic Potential Analysis
            print("⚡ Analyzing electrostatic potential...")
            try:
                charged_residues = {'ARG': 1, 'LYS': 1, 'HIS': 0.5, 'ASP': -1, 'GLU': -1}
                electrostatic_score = 0
                nearby_charges = 0
                
                for model in structure:
                    for chain in model:
                        for residue in chain:
                            if residue.get_id()[0] == ' ':  # Protein residue
                                resname = residue.get_resname()
                                if resname in charged_residues:
                                    # Check if residue is near the pocket center
                                    try:
                                        ca_atom = residue['CA']
                                        dist = np.linalg.norm(ca_atom.get_coord() - center_coords)
                                        if dist <= 10.0:  # Within 10Å of pocket center
                                            electrostatic_score += charged_residues[resname]
                                            nearby_charges += 1
                                    except:
                                        continue
                
                results['electrostatic_score'] = electrostatic_score
                results['nearby_charged_residues'] = nearby_charges
                print(f"✓ Electrostatic analysis completed (score: {electrostatic_score})")
                
            except Exception as e:
                print(f"⚠️  Electrostatic analysis failed: {e}")
                results['electrostatic_score'] = 'N/A'
            
            # 3. Hydrophobic Potential Analysis
            print("💧 Analyzing hydrophobic potential...")
            try:
                hydrophobic_residues = ['ALA', 'VAL', 'LEU', 'ILE', 'MET', 'PHE', 'TRP', 'TYR', 'PRO']
                hydrophobic_score = 0
                nearby_hydrophobic = 0
                
                for model in structure:
                    for chain in model:
                        for residue in chain:
                            if residue.get_id()[0] == ' ':  # Protein residue
                                resname = residue.get_resname()
                                if resname in hydrophobic_residues:
                                    try:
                                        ca_atom = residue['CA']
                                        dist = np.linalg.norm(ca_atom.get_coord() - center_coords)
                                        if dist <= 8.0:  # Within 8Å of pocket center
                                            hydrophobic_score += 1
                                            nearby_hydrophobic += 1
                                    except:
                                        continue
                
                results['hydrophobic_score'] = hydrophobic_score
                results['nearby_hydrophobic_residues'] = nearby_hydrophobic
                print(f"✓ Hydrophobic analysis completed (score: {hydrophobic_score})")
                
            except Exception as e:
                print(f"⚠️  Hydrophobic analysis failed: {e}")
                results['hydrophobic_score'] = 'N/A'
            
            # 4. Druggability Prediction
            print("💊 Calculating druggability score...")
            try:
                druggability_score = 0.0
                
                # Volume contribution
                if isinstance(results.get('pocket_volume_A3'), (int, float)):
                    if results['pocket_volume_A3'] > 200:
                        druggability_score += 0.3
                    elif results['pocket_volume_A3'] > 100:
                        druggability_score += 0.2
                
                # Hydrophobic contribution
                if isinstance(results.get('hydrophobic_score'), (int, float)):
                    if results['hydrophobic_score'] > 5:
                        druggability_score += 0.4
                    elif results['hydrophobic_score'] > 3:
                        druggability_score += 0.3
                
                # Balance of charged residues
                if isinstance(results.get('electrostatic_score'), (int, float)):
                    abs_charge = abs(results['electrostatic_score'])
                    if abs_charge <= 2:
                        druggability_score += 0.3
                    elif abs_charge <= 4:
                        druggability_score += 0.2
                
                results['druggability_score'] = min(druggability_score, 1.0)
                print(f"✓ Druggability score: {results['druggability_score']:.2f}")
                
            except Exception as e:
                print(f"⚠️  Druggability calculation failed: {e}")
                results['druggability_score'] = 'N/A'
            
            return results
            
        except Exception as e:
            raise ValueError(f"Pocket analysis failed: {e}")

    def generate_summary_report(self, results: Dict[str, Any], pdb_id: str, excel_workbook: Optional[Workbook] = None, sheet_name: Optional[str] = None):
        """
        Generate comprehensive summary report with Excel support.
        
        Args:
            results (Dict[str, Any]): Pipeline results dictionary
            pdb_id (str): PDB identifier
            excel_workbook (Workbook, optional): Excel workbook for multi-PDB analysis
            sheet_name (str, optional): Excel sheet name
        """
        print("\n📊 Generating Summary Report...")
        
        # Create DataFrame for easy CSV export
        summary_data = []
        
        # Basic information
        summary_data.append(['PDB_ID', pdb_id])
        summary_data.append(['Selected_Ligand', results.get('selected_ligand', 'N/A')])
        summary_data.append(['Active_Site_Center_X', f"{results.get('center_x', 0):.3f}"])
        summary_data.append(['Active_Site_Center_Y', f"{results.get('center_y', 0):.3f}"])
        summary_data.append(['Active_Site_Center_Z', f"{results.get('center_z', 0):.3f}"])
        summary_data.append(['Interacting_Atoms_Count', results.get('num_interacting_atoms', 'N/A')])
        
        # Pocket properties
        summary_data.append(['Pocket_Volume_A3', results.get('pocket_volume_A3', 'N/A')])
        summary_data.append(['Electrostatic_Score', results.get('electrostatic_score', 'N/A')])
        summary_data.append(['Nearby_Charged_Residues', results.get('nearby_charged_residues', 'N/A')])
        summary_data.append(['Hydrophobic_Score', results.get('hydrophobic_score', 'N/A')])
        summary_data.append(['Nearby_Hydrophobic_Residues', results.get('nearby_hydrophobic_residues', 'N/A')])
        summary_data.append(['Druggability_Score', results.get('druggability_score', 'N/A')])
        
        # File paths
        summary_data.append(['Original_PDB', results.get('original_pdb', 'N/A')])
        summary_data.append(['Cleaned_PDB', results.get('cleaned_pdb', 'N/A')])
        summary_data.append(['Ligand_PDB', results.get('ligand_pdb', 'N/A')])
        
        # Save as CSV
        df = pd.DataFrame(summary_data, columns=['Property', 'Value'])
        csv_filename = self.output_dir / f"{pdb_id}_pipeline_results.csv"
        df.to_csv(csv_filename, index=False)
        
        # Save to Excel if available
        if EXCEL_AVAILABLE and excel_workbook is not None:
            self._add_to_excel(excel_workbook, sheet_name or "Results", summary_data, pdb_id)
        
        # Display summary
        print("\n📋 Pipeline Summary:")
        print("-" * 40)
        for prop, value in summary_data[:12]:  # Show main results
            print(f"{prop:<25}: {value}")
        
        print(f"\n✓ Detailed results saved to: {csv_filename}")
        
        if EXCEL_AVAILABLE and excel_workbook is not None:
            excel_filename = self.output_dir / "multi_pdb_analysis.xlsx"
            excel_workbook.save(str(excel_filename))
            print(f"✓ Excel results saved to: {excel_filename}")
    
    def _add_to_excel(self, workbook: Workbook, sheet_name: str, summary_data: List[List], pdb_id: str):
        """
        Add results to Excel workbook.
        
        Args:
            workbook (Workbook): Excel workbook
            sheet_name (str): Sheet name
            summary_data (List[List]): Summary data
            pdb_id (str): PDB identifier
        """
        try:
            # Get or create worksheet
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
            else:
                ws = workbook.create_sheet(sheet_name)
                # Add headers
                headers = ['PDB_ID', 'Property', 'Value']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Find next empty row
            next_row = ws.max_row + 1
            
            # Add data
            for prop, value in summary_data:
                ws.cell(row=next_row, column=1, value=pdb_id)
                ws.cell(row=next_row, column=2, value=prop)
                ws.cell(row=next_row, column=3, value=value)
                next_row += 1
            
            # Add separator row
            ws.cell(row=next_row, column=1, value="-" * 20)
            ws.cell(row=next_row, column=2, value="-" * 20)
            ws.cell(row=next_row, column=3, value="-" * 20)
            
        except Exception as e:
            print(f"⚠️  Excel writing failed: {e}")
    
    def create_excel_workbook(self) -> Optional[Workbook]:
        """
        Create Excel workbook for multi-PDB analysis.
        
        Returns:
            Optional[Workbook]: Excel workbook or None if not available
        """
        if not EXCEL_AVAILABLE:
            return None
        
        try:
            workbook = Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)
            return workbook
        except Exception as e:
            print(f"⚠️  Failed to create Excel workbook: {e}")
            return None

    def run_complete_pipeline(self, pdb_id: str, interactive: bool = True) -> Dict[str, Any]:
        """
        Run the complete molecular docking pipeline.
        
        Args:
            pdb_id (str): PDB identifier
            interactive (bool): Whether to run in interactive mode
            
        Returns:
            Dict[str, Any]: Complete pipeline results
        """
        print("🚀 Starting Complete Molecular Docking Pipeline")
        print("=" * 50)
        
        results = {}
        
        try:
            # Step 1: Fetch PDB
            pdb_file = self.fetch_pdb(pdb_id)
            results['original_pdb'] = pdb_file
            
            # Step 2: Chain Management (optional)
            if interactive:
                manage_chains_choice = input("Do you want to filter chains? (y/n): ").lower().strip()
                if manage_chains_choice == 'y':
                    pdb_file = self.manage_chains(pdb_file)
                    results['chain_filtered_pdb'] = pdb_file
            
            # Step 3: Enumerate HETATMs
            hetatm_details, unique_hetatms = self.enumerate_hetatms(pdb_file)
            
            if not hetatm_details:
                raise ValueError("No HETATMs found in the structure")
            
            # Step 4: Select ligand interactively
            if interactive:
                # Group HETATMs by type for easier selection
                hetatm_groups = {}
                for resname, chain_id, res_id, residue in hetatm_details:
                    if resname not in hetatm_groups:
                        hetatm_groups[resname] = []
                    hetatm_groups[resname].append((resname, chain_id, res_id, residue))
                
                print("\n📋 Available HETATM Groups:")
                group_choices = []
                choice_num = 1
                
                for resname, instances in sorted(hetatm_groups.items()):
                    print(f"   {choice_num}. {resname} ({len(instances)} instance(s))")
                    group_choices.append((resname, instances))
                    choice_num += 1
                
                while True:
                    try:
                        group_choice = int(input(f"Select HETATM type (1-{len(group_choices)}): ")) - 1
                        if 0 <= group_choice < len(group_choices):
                            selected_resname, instances = group_choices[group_choice]
                            
                            if len(instances) == 1:
                                # Only one instance, select it automatically
                                selected_hetatm, chain_id, res_id, residue = instances[0]
                                print(f"✓ Selected: {selected_hetatm}_{chain_id}_{res_id}")
                                break
                            else:
                                # Multiple instances, let user choose specific one
                                print(f"\n📋 Available {selected_resname} instances:")
                                for i, (resname, chain_id, res_id, residue) in enumerate(instances):
                                    print(f"   {i+1}. {resname}_{chain_id}_{res_id}")
                                
                                instance_choice = int(input(f"Select instance (1-{len(instances)}): ")) - 1
                                if 0 <= instance_choice < len(instances):
                                    selected_hetatm, chain_id, res_id, residue = instances[instance_choice]
                                    print(f"✓ Selected: {selected_hetatm}_{chain_id}_{res_id}")
                                    break
                                else:
                                    print("Invalid choice, please try again.")
                        else:
                            print("Invalid choice, please try again.")
                    except ValueError:
                        print("Please enter a valid number.")
            else:
                # For non-interactive mode, select first HETATM
                selected_hetatm, chain_id, res_id, residue = hetatm_details[0]
            
            print(f"✓ Selected ligand: {selected_hetatm}_{chain_id}_{res_id}")
            results['selected_ligand'] = f"{selected_hetatm}_{chain_id}_{res_id}"
            
            # Store selected ligand info for cleaning warnings
            self.selected_ligand_info = {
                'resname': selected_hetatm,
                'chain_id': chain_id,
                'res_id': res_id
            }
            
            # Step 5: Save ligand as separate PDB
            ligand_pdb = self.save_hetatm_as_pdb(pdb_file, selected_hetatm, chain_id, res_id)
            results['ligand_pdb'] = ligand_pdb
            
            # Step 6: Clean PDB
            if interactive:
                cleaned_pdb = self.clean_pdb(pdb_file)
            else:
                # Default cleaning - remove water and common ions
                cleaned_pdb = self.clean_pdb(pdb_file, to_remove_list=['HOH', 'NA', 'CL'])
            results['cleaned_pdb'] = cleaned_pdb
            
            # Step 7: Extract active site coordinates
            try:
                # Check if the selected ligand still exists in the cleaned structure
                parser = PDBParser(QUIET=True)
                cleaned_structure = parser.get_structure('protein', cleaned_pdb)
                
                ligand_found = False
                for model in cleaned_structure:
                    for chain in model:
                        for residue in chain:
                            if (residue.get_resname() == selected_hetatm and 
                                residue.get_id()[1] == res_id and
                                residue.get_id()[0] != ' '):
                                ligand_found = True
                                break
                
                if not ligand_found:
                    print(f"⚠️  Selected ligand {selected_hetatm}_{chain_id}_{res_id} was removed during cleaning")
                    print("🔄 Using original structure for active site analysis...")
                    # Use original structure for analysis
                    avg_coords, num_atoms = self.extract_active_site_coords(
                        pdb_file, selected_hetatm, chain_id, res_id, method='distance'
                    )
                else:
                    # Use cleaned structure for analysis
                    avg_coords, num_atoms = self.extract_active_site_coords(
                        cleaned_pdb, selected_hetatm, chain_id, res_id, method='distance'
                    )
                
                results['active_site_center'] = avg_coords.tolist()
                results['num_interacting_atoms'] = num_atoms
            except Exception as e:
                print(f"❌ Failed to extract active site coordinates: {e}")
                return results
            
            # Step 8: Comprehensive pocket analysis
            pocket_results = self.analyze_pocket_properties(
                cleaned_pdb, avg_coords, ligand_pdb
            )
            results.update(pocket_results)
            
            # Step 9: Generate summary report
            self.generate_summary_report(results, pdb_id)
            
            print("\n🎉 Pipeline completed successfully!")
            print(f"📁 All files saved in: {self.output_dir}")
            
            return results
            
        except Exception as e:
            print(f"❌ Pipeline failed: {e}")
            return results


def main():
    """Main function to run the pipeline with multi-PDB support."""
    print("🔬 PDB Prepare Wizard - Molecular Docking Pipeline")
    print("=" * 50)
    
    # Initialize pipeline
    pipeline = MolecularDockingPipeline()
    
    # Ask if user wants to analyze multiple PDBs
    multi_analysis = input("Do you want to analyze multiple PDBs? (y/n): ").lower().strip()
    
    excel_workbook = None
    if multi_analysis == 'y':
        if EXCEL_AVAILABLE:
            excel_workbook = pipeline.create_excel_workbook()
            print("✓ Excel workbook created for multi-PDB analysis")
        else:
            print("⚠️  Excel support not available - will use CSV files only")
    
    pdb_count = 0
    while True:
        # Get PDB ID from user
        pdb_id = input(f"\nEnter PDB ID #{pdb_count + 1} (e.g., 1ABC) or 'quit' to exit: ").strip().upper()
        
        if pdb_id.lower() == 'quit':
            break
            
        if not pdb_id or len(pdb_id) != 4:
            print("❌ Invalid PDB ID. Please enter a 4-character PDB identifier.")
            continue
        
        try:
            # Run pipeline
            results = pipeline.run_complete_pipeline(pdb_id, interactive=True)
            
            if results:
                # Generate report with Excel support
                pipeline.generate_summary_report(results, pdb_id, excel_workbook)
                pdb_count += 1
                print(f"\n✅ PDB {pdb_id} analysis completed successfully!")
            else:
                print(f"\n❌ PDB {pdb_id} analysis failed. Please check the error messages above.")
                
        except Exception as e:
            print(f"\n❌ Error analyzing PDB {pdb_id}: {e}")
        
        # Ask if user wants to continue
        if multi_analysis == 'y':
            continue_choice = input("\nDo you want to analyze another PDB? (y/n): ").lower().strip()
            if continue_choice != 'y':
                break
        else:
            break
    
    # Final summary
    if pdb_count > 0:
        print(f"\n🎉 Analysis completed! Processed {pdb_count} PDB structure(s).")
        if excel_workbook and EXCEL_AVAILABLE:
            excel_filename = pipeline.output_dir / "multi_pdb_analysis.xlsx"
            excel_workbook.save(str(excel_filename))
            print(f"📊 All results saved to: {excel_filename}")
    else:
        print("\n❌ No PDB structures were successfully analyzed.")


if __name__ == "__main__":
    main() 