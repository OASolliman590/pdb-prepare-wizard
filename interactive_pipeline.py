#!/usr/bin/env python3
"""
PDB Prepare Wizard - Interactive Pipeline
========================================

Interactive version of the PDB preparation pipeline with user-friendly prompts
and multi-PDB analysis capabilities.

Author: Molecular Docking Pipeline
Version: 2.1.0
"""

import os
import sys
from pathlib import Path
from typing import Optional

# Import the core pipeline
from core_pipeline import MolecularDockingPipeline, EXCEL_AVAILABLE

if EXCEL_AVAILABLE:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

def get_user_choice(prompt: str, valid_choices: list, default: Optional[str] = None) -> str:
    """Get user choice with validation."""
    while True:
        if default:
            user_input = input(f"{prompt} (default: {default}): ").strip()
            if not user_input:
                return default
        else:
            user_input = input(f"{prompt}: ").strip()
        
        if user_input.lower() in [choice.lower() for choice in valid_choices]:
            return user_input.lower()
        
        print(f"❌ Invalid choice. Please select from: {', '.join(valid_choices)}")

def select_hetatm_interactive(hetatm_details, unique_hetatms):
    """Interactive HETATM selection with grouped display."""
    if not unique_hetatms:
        print("❌ No HETATMs found to select from")
        return None, None, None
    
    print("\n📋 Available HETATM types:")
    hetatm_counts = {}
    for resname, _, _, _ in hetatm_details:
        hetatm_counts[resname] = hetatm_counts.get(resname, 0) + 1
    
    for i, (resname, count) in enumerate(hetatm_counts.items(), 1):
        print(f"   {i}. {resname} ({count} instance(s))")
    
    # Get user selection by type
    while True:
        try:
            choice = input(f"\nSelect HETATM type (1-{len(hetatm_counts)}): ").strip()
            choice_num = int(choice) - 1
            if 0 <= choice_num < len(hetatm_counts):
                selected_type = list(hetatm_counts.keys())[choice_num]
                break
            else:
                print(f"❌ Please enter a number between 1 and {len(hetatm_counts)}")
        except ValueError:
            print("❌ Please enter a valid number")
    
    # Find instances of selected type
    instances = [(chain_id, res_id, residue) for resname, chain_id, res_id, residue 
                in hetatm_details if resname == selected_type]
    
    if len(instances) == 1:
        chain_id, res_id, residue = instances[0]
        print(f"✓ Selected: {selected_type} in chain {chain_id}, residue {res_id}")
        return selected_type, chain_id, res_id
    else:
        print(f"\n📋 Multiple instances of {selected_type} found:")
        for i, (chain_id, res_id, _) in enumerate(instances, 1):
            print(f"   {i}. Chain {chain_id}, Residue {res_id}")
        
        while True:
            try:
                choice = input(f"Select instance (1-{len(instances)}): ").strip()
                choice_num = int(choice) - 1
                if 0 <= choice_num < len(instances):
                    chain_id, res_id, _ = instances[choice_num]
                    print(f"✓ Selected: {selected_type} in chain {chain_id}, residue {res_id}")
                    return selected_type, chain_id, res_id
                else:
                    print(f"❌ Please enter a number between 1 and {len(instances)}")
            except ValueError:
                print("❌ Please enter a valid number")

def get_cleaning_choice(unique_hetatms):
    """Interactive cleaning choice with smart options."""
    print("\n📝 Cleaning Options:")
    print("   1. Remove specific residues (enter comma-separated list)")
    print("   2. Remove ALL HETATMs (enter 'ALL')")
    print("   3. Remove common residues only (enter 'COMMON')")
    print("   4. Skip cleaning (enter 'SKIP')")
    
    choice = input("\nEnter your choice (1-4, comma-separated list, ALL, COMMON, or SKIP): ").strip().upper()
    
    if choice == 'ALL':
        return unique_hetatms
    elif choice == 'COMMON':
        common_residues = ['HOH', 'NA', 'CL', 'SO4', 'CA', 'MG', 'ZN', 'FE', 'CU', 'MN']
        return [r for r in common_residues if r in unique_hetatms]
    elif choice == 'SKIP':
        return []
    elif choice in ['1', '2', '3', '4']:
        if choice == '1':
            residues_input = input("Enter comma-separated residue names to remove: ").strip()
            return [r.strip().upper() for r in residues_input.split(',') if r.strip()]
        elif choice == '2':
            return unique_hetatms
        elif choice == '3':
            common_residues = ['HOH', 'NA', 'CL', 'SO4', 'CA', 'MG', 'ZN', 'FE', 'CU', 'MN']
            return [r for r in common_residues if r in unique_hetatms]
        else:  # choice == '4'
            return []
    else:
        # Assume comma-separated list
        return [r.strip().upper() for r in choice.split(',') if r.strip()]

def show_removal_summary(pdb_file, to_remove_list):
    """Show what will be removed before confirmation."""
    if not to_remove_list:
        print("✓ No residues will be removed")
        return True
    
    try:
        from Bio.PDB import PDBParser
        parser = PDBParser(QUIET=True)
        structure = parser.get_structure('protein', pdb_file)
        
        removal_counts = {}
        for model in structure:
            for chain in model:
                for residue in chain:
                    resname = residue.get_resname()
                    if resname in to_remove_list:
                        removal_counts[resname] = removal_counts.get(resname, 0) + 1
        
        if not removal_counts:
            print("✓ No matching residues found to remove")
            return True
        
        print(f"\n📊 Removal Summary:")
        total_removed = sum(removal_counts.values())
        print(f"   Total residues to remove: {total_removed}")
        for resname, count in removal_counts.items():
            print(f"   - {resname}: {count} residues")
        
        confirm = get_user_choice("\nProceed with removal?", ['y', 'n'], 'y')
        return confirm == 'y'
        
    except Exception as e:
        print(f"⚠️  Could not generate removal summary: {e}")
        confirm = get_user_choice("Proceed with removal anyway?", ['y', 'n'], 'n')
        return confirm == 'y'

def run_single_pdb_analysis(pipeline, pdb_id, excel_workbook=None):
    """Run analysis for a single PDB."""
    print(f"\n🚀 Starting analysis for PDB {pdb_id}")
    print("=" * 50)
    
    try:
        # Step 1: Download PDB
        pdb_file = pipeline.fetch_pdb(pdb_id)
        
        # Step 2: Enumerate HETATMs
        hetatm_details, unique_hetatms = pipeline.enumerate_hetatms(pdb_file)
        if not unique_hetatms:
            print("⚠️  No HETATMs found. Proceeding with protein-only analysis.")
            selected_hetatm = None
            chain_id = None
            res_id = None
        else:
            # Step 3: Select ligand
            selected_hetatm, chain_id, res_id = select_hetatm_interactive(hetatm_details, unique_hetatms)
            
            if selected_hetatm:
                # Save ligand as separate PDB
                ligand_pdb = pipeline.save_hetatm_as_pdb(
                    pdb_file, selected_hetatm, chain_id, res_id
                )
        
        # Step 4: Extract active site coordinates BEFORE cleaning (if ligand selected)
        # This allows us to get binding site info even if we remove the ligand later
        results = {'pdb_id': pdb_id}
        coords = None
        num_atoms = 0
        
        if selected_hetatm:
            try:
                print("\n🔄 Extracting active site coordinates from original structure...")
                coords, num_atoms = pipeline.extract_active_site_coords(
                    pdb_file, selected_hetatm, chain_id, res_id  # Use original PDB with ligand
                )
                results.update({
                    'selected_ligand': f"{selected_hetatm}_{chain_id}_{res_id}",
                    'active_site_center_x': coords[0],
                    'active_site_center_y': coords[1],
                    'active_site_center_z': coords[2],
                    'interacting_atoms_count': num_atoms
                })
                print(f"✓ Active site coordinates extracted: X={coords[0]:.2f}, Y={coords[1]:.2f}, Z={coords[2]:.2f}")
            except Exception as e:
                print(f"⚠️  Active site analysis failed: {e}")
                print("Proceeding without active site coordinates.")
        
        # Step 5: Clean PDB (remove ligand and other unwanted residues)
        to_remove_list = get_cleaning_choice(unique_hetatms)
        
        if selected_hetatm and selected_hetatm in to_remove_list:
            print(f"\n⚠️  NOTE: You selected {selected_hetatm} as your ligand.")
            print("   Active site coordinates have been extracted from the original structure.")
            print("   Removing the ligand will create an apo (ligand-free) receptor for docking.")
            keep_ligand = get_user_choice("Remove ligand from cleaned structure? (y=apo receptor, n=keep ligand)", ['y', 'n'], 'y')
            if keep_ligand == 'n':
                to_remove_list.remove(selected_hetatm)
        
        if show_removal_summary(pdb_file, to_remove_list):
            cleaned_pdb = pipeline.clean_pdb(pdb_file, to_remove_list)
        else:
            print("Cleaning cancelled. Using original structure.")
            cleaned_pdb = pdb_file
        
        # Step 6: Analyze pocket properties using cleaned structure (if coordinates were extracted)
        if selected_hetatm and coords is not None:
            try:
                # Use the coordinates extracted from original structure, but analyze cleaned structure
                pocket_results = pipeline.analyze_pocket_properties(cleaned_pdb, coords)
                results.update(pocket_results)
            except Exception as e:
                print(f"⚠️  Pocket analysis failed: {e}")
                print("Proceeding without pocket analysis.")
        
        # Step 7: Generate reports
        report_file = pipeline.generate_summary_report(results, pdb_id)
        
        # Add to Excel workbook if provided
        if excel_workbook and EXCEL_AVAILABLE:
            add_to_excel_workbook(excel_workbook, pdb_id, results)
        
        print(f"\n✅ Analysis completed successfully for PDB {pdb_id}!")
        return True
        
    except Exception as e:
        print(f"\n❌ Analysis failed for PDB {pdb_id}: {e}")
        return False

def add_to_excel_workbook(workbook, pdb_id, results):
    """Add results to Excel workbook."""
    try:
        # Create or get worksheet
        if f"Summary" not in workbook.sheetnames:
            ws = workbook.create_sheet("Summary")
            # Add headers
            headers = ["PDB_ID", "Property", "Value"]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        else:
            ws = workbook["Summary"]
        
        # Add results
        for key, value in results.items():
            ws.append([pdb_id, key, value])
            
        # Add separator row
        ws.append(["", "", ""])
        
    except Exception as e:
        print(f"⚠️  Failed to add results to Excel: {e}")

def run_interactive_pipeline(output_dir: Optional[str] = None):
    """Run the interactive pipeline.
    
    Args:
        output_dir: Optional output directory path. If None, user will be prompted.
    """
    print("🔬 PDB Prepare Wizard - Interactive Pipeline")
    print("=" * 50)
    
    # Initialize pipeline
    if output_dir is None:
        output_dir = input("Enter output directory (default: pipeline_output): ").strip()
        if not output_dir:
            output_dir = "pipeline_output"
    else:
        print(f"✓ Using output directory: {output_dir}")
    
    pipeline = MolecularDockingPipeline(output_dir)
    
    # Check if user wants multi-PDB analysis
    multi_pdb = get_user_choice("Do you want to analyze multiple PDBs?", ['y', 'n'], 'n')
    
    excel_workbook = None
    if multi_pdb == 'y' and EXCEL_AVAILABLE:
        excel_workbook = Workbook()
        # Remove default sheet
        if "Sheet" in excel_workbook.sheetnames:
            excel_workbook.remove(excel_workbook["Sheet"])
        print("✓ Excel workbook created for multi-PDB analysis")
    
    pdb_count = 0
    while True:
        pdb_count += 1
        
        if multi_pdb == 'y':
            pdb_id = input(f"\nEnter PDB ID #{pdb_count} (e.g., 1ABC) or 'quit' to exit: ").strip().upper()
        else:
            pdb_id = input("\nEnter PDB ID (e.g., 1ABC): ").strip().upper()
        
        if pdb_id.lower() == 'quit':
            break
        
        if not pdb_id:
            print("❌ Please enter a valid PDB ID")
            pdb_count -= 1
            continue
        
        # Validate PDB ID format
        if len(pdb_id) != 4 or not pdb_id.isalnum():
            print("❌ PDB ID should be 4 characters (letters and numbers)")
            pdb_count -= 1
            continue
        
        # Run analysis
        success = run_single_pdb_analysis(pipeline, pdb_id, excel_workbook)
        
        if multi_pdb == 'y':
            if success:
                continue_analysis = get_user_choice("Do you want to analyze another PDB?", ['y', 'n'], 'n')
                if continue_analysis == 'n':
                    break
            else:
                continue_anyway = get_user_choice("Continue with next PDB despite this failure?", ['y', 'n'], 'y')
                if continue_anyway == 'n':
                    break
        else:
            break
    
    # Save Excel file for multi-PDB analysis
    if excel_workbook and EXCEL_AVAILABLE and multi_pdb == 'y':
        excel_path = Path(output_dir) / "multi_pdb_analysis.xlsx"
        excel_workbook.save(excel_path)
        print(f"\n📊 All results saved to: {excel_path}")
    
    if multi_pdb == 'y':
        print(f"\n🎉 Analysis completed! Processed {pdb_count - 1} PDB structure(s).")
    
    print("\n✓ Pipeline execution completed!")

if __name__ == "__main__":
    try:
        run_interactive_pipeline()
    except KeyboardInterrupt:
        print("\n\n⚠️  Analysis interrupted by user")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        sys.exit(1)

